"""Classeur Excel : un onglet par vue, en-têtes gras, filtres et volets figés.

`openpyxl` est déjà présent (dépendance de DocFuse) ; aucun ajout requis.

Le classeur est produit en **mode écriture seule** (`Workbook(write_only=True)`) :
chaque ligne part vers le fichier dès qu'elle est produite, openpyxl ne conserve
aucune cellule en mémoire, et l'onglet « Fichiers » est alimenté directement par
le curseur SQLite de `Database.latest_analyses` — jamais par une liste complète.
Une campagne de 934 000 fichiers passait de 12 Go de mémoire à quelques centaines
de mégaoctets.

Deux conséquences, assumées et documentées :

- **Largeurs de colonnes** : l'élément `cols` du XML précède les lignes, les
  largeurs doivent donc être connues *avant* la première écriture. Elles sont
  dérivées d'un échantillon borné — les `WIDTH_SAMPLE_ROWS` premières lignes de
  l'onglet — et non de la totalité : une valeur exceptionnellement longue en fin
  d'onglet n'élargit plus sa colonne.
- **Limite d'Excel** : `MAX_SHEET_ROWS` (1 048 576) lignes par feuille, en-tête
  comprise. Au-delà, l'onglet est **tronqué** et sa dernière ligne porte un
  avertissement visible ; le même message est remis à `on_warning` (par défaut le
  journal, que la CLI affiche). Les formats `powerbi` et `csv` n'ont pas cette
  limite et restent la sortie complète.

Ce que le mode écriture seule permet encore, et qui est conservé : volets figés
(`freeze_panes`), filtre automatique (`auto_filter`, écrit à la fermeture de
l'onglet), styles de cellule (`WriteOnlyCell`) et formats de nombre. Ce qu'il
interdit : relire ou modifier une cellule déjà écrite.

Deux détails de forme du fichier changent, sans effet pour Excel : les textes
sont écrits en ligne (`inlineStr`) plutôt que dans la table de chaînes partagées
— c'est justement ce qui évitait de tout garder en mémoire — et l'élément
`dimension` (une simple indication de l'étendue utilisée) n'est plus écrit,
puisqu'il précède les lignes et que leur nombre n'est pas encore connu.
"""

from __future__ import annotations

import logging
from collections.abc import Callable, Iterable, Iterator, Sequence
from dataclasses import dataclass
from datetime import date
from itertools import chain, islice
from pathlib import Path
from typing import Any

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.cell import WriteOnlyCell  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from docia import views
from docia.db import Database
from docia.report import tabular
from docia.report.data import ReportData, collect

logger = logging.getLogger(__name__)

SCOPE_SHEET = "Périmètre"
"""Onglet ajouté **en premier**, et seulement quand l'inventaire est incomplet.

Le classeur sert aux mêmes décisions de suppression que le rapport HTML : quand
une cible n'a pas été parcourue ou qu'un scan a été arrêté, cela doit se voir à
l'ouverture, pas au détour d'une ligne de la synthèse. Sur une campagne complète
— le cas normal — le classeur garde exactement les onglets de `SHEETS`."""

SHEETS: tuple[str, ...] = (
    "Synthèse",
    "Fichiers",
    "Doublons",
    "Ancienneté",
    "Sensibles",
    "Conservation",
    "Nettoyage",
    "Revues",
    "Erreurs",
)
"""Onglets du classeur, dans l'ordre (`SCOPE_SHEET` s'ajoute devant si besoin)."""

MAX_SHEET_ROWS = 1_048_576
"""Lignes par feuille acceptées par Excel (en-tête comprise) — limite du format xlsx."""

WIDTH_SAMPLE_ROWS = 200
"""Lignes examinées pour dimensionner les colonnes (voir le docstring du module)."""

_HEADER_FILL = PatternFill("solid", fgColor="D9E2EC")
_HEADER_FONT = Font(bold=True, color="1F2933")
_WARNING_FILL = PatternFill("solid", fgColor="FFE066")
_WARNING_FONT = Font(bold=True, color="9B1C1C")
_MAX_WIDTH = 70

Row = Sequence[Any]
"""Une ligne de données : les valeurs, dans l'ordre des en-têtes."""


def _grouped(value: int) -> str:
    """`1048576` → `1 048 576` (espaces, comme partout dans l'outil)."""
    return f"{value:,}".replace(",", " ")


def _truncation_message(title: str, written: int, dropped: int) -> str:
    """Message rendu à l'utilisateur *et* écrit dans l'onglet tronqué."""
    return (
        f"Onglet « {title} » tronqué : {_grouped(written)} ligne(s) écrite(s), "
        f"{_grouped(dropped)} non écrite(s) — Excel n'accepte que "
        f"{_grouped(MAX_SHEET_ROWS)} lignes par feuille. Pour la totalité, utiliser "
        "« export --format powerbi » ou « export --format csv », qui n'ont pas cette limite."
    )


def _header_cells(sheet: Any, headers: Sequence[str]) -> list[Any]:
    cells = []
    for text in headers:
        cell = WriteOnlyCell(sheet, value=text)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cells.append(cell)
    return cells


def _data_cells(sheet: Any, values: Row, formats: dict[int, str]) -> list[Any]:
    """Ligne prête à écrire : valeurs brutes, sauf format de nombre ou texte à risque.

    openpyxl décide seul du type d'une valeur : une chaîne commençant par `=`
    devient une **formule** (`<f>…</f>` dans le XML), pas du texte. Le nom d'un
    fichier du partage et surtout le `resume` et les justifications de la LLM
    passent par ici ; ils sont donc réécrits en chaîne (`data_type = "s"`,
    c'est-à-dire `inlineStr`) — voir `docia.report.tabular`. Les nombres et les
    dates ne sont pas touchés : ils doivent rester calculables.
    """
    row: list[Any] = []
    for index, value in enumerate(values, start=1):
        number_format = formats.get(index)
        if number_format is None and not tabular.is_formula_text(value):
            row.append(value)
            continue
        cell = WriteOnlyCell(sheet, value=value)
        if number_format is not None:
            cell.number_format = number_format
        if cell.data_type == "f":
            cell.data_type = "s"  # texte, jamais formule
        row.append(cell)
    return row


def _note_cell(sheet: Any, text: str) -> Any:
    """Ligne d'information en bas d'onglet, dans le style des avertissements."""
    cell = WriteOnlyCell(sheet, value=text)
    cell.font = _WARNING_FONT
    cell.fill = _WARNING_FILL
    return cell


def _add_sheet(
    workbook: Workbook,
    title: str,
    headers: Sequence[str],
    rows: Iterable[Row],
    *,
    formats: dict[int, str] | None = None,
    max_rows: int = MAX_SHEET_ROWS,
    note: str | None = None,
) -> str | None:
    """Ajoute un onglet écrit en flux ; rend le message de troncature, ou `None`.

    `rows` n'est parcouru qu'une fois et n'est jamais matérialisé : seules les
    `WIDTH_SAMPLE_ROWS` premières lignes sont retenues, le temps de dimensionner
    les colonnes. `max_rows` compte l'en-tête ; abaissable pour les tests.

    `note` est une dernière ligne d'information — « ce tableau est un classement
    borné, voici sur combien il porte et où est la liste entière ». Elle prend une
    ligne sur le budget de l'onglet, comme l'avertissement de troncature.
    """
    sheet = workbook.create_sheet(title=title[:31])
    number_formats = formats or {}
    source = iter(rows)
    sample = list(islice(source, WIDTH_SAMPLE_ROWS))
    if note is not None:
        max_rows = max(max_rows - 1, 1)

    widths = [len(h) + 2 for h in headers]
    for row in sample:
        for index, value in enumerate(row):
            if index < len(widths):
                widths[index] = max(widths[index], min(len(str(value)) + 2, _MAX_WIDTH))
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.append(_header_cells(sheet, headers))

    limit = max(max_rows - 1, 0)  # lignes de données possibles, en-tête déduite
    written = dropped = 0
    held: Row | None = None
    for row in chain(sample, source):
        if written < limit - 1:
            sheet.append(_data_cells(sheet, row, number_formats))
            written += 1
        elif held is None and limit > 0:
            held = row  # dernière place : donnée si le flux s'arrête là, avertissement sinon
        else:
            dropped += 1

    message: str | None = None
    if dropped:
        if held is not None:
            dropped += 1  # la ligne retenue cède sa place à l'avertissement
            held = None
        # `limit == 0` : pas même une ligne pour l'avertissement. Il ne serait
        # sinon compté qu'une ligne perdue de plus qu'il n'y en a vraiment, et
        # l'onglet dépasserait `max_rows` d'une ligne.
        message = _truncation_message(title, written, dropped)
        if limit > 0:
            warning = WriteOnlyCell(sheet, value=message)
            warning.font = _WARNING_FONT
            warning.fill = _WARNING_FILL
            sheet.append([warning])
    elif held is not None:
        sheet.append(_data_cells(sheet, held, number_formats))
        written += 1

    extra = 1 if message and limit > 0 else 0
    if note is not None:
        sheet.append([_note_cell(sheet, note)])
        extra += 1

    if headers:
        last = get_column_letter(len(headers))
        total = written + extra + 1
        sheet.auto_filter.ref = f"A1:{last}{max(total, 2)}"
    return message


def _date_cell(value: date | None) -> date | str:
    return value if value is not None else ""


def _file_rows(db: Database) -> tuple[list[str], Iterator[Row]]:
    """En-têtes et lignes de l'onglet « Fichiers », lus au fil du curseur SQLite.

    La première ligne est consommée pour connaître les colonnes, puis remise en
    tête du flux : la campagne entière n'est jamais en mémoire.
    """
    cursor = db.latest_analyses()
    first = next(cursor, None)
    headers = list(first.keys()) if first is not None else ["path"]

    def rows() -> Iterator[Row]:
        if first is None:
            return
        for record in chain([first], cursor):
            yield [record[name] if record[name] is not None else "" for name in headers]

    return headers, rows()


@dataclass(frozen=True)
class _Sheet:
    """Un onglet à écrire : titre, en-têtes, lignes (flux), formats de colonnes, note finale."""

    title: str
    headers: Sequence[str]
    rows: Iterable[Row]
    formats: dict[int, str] | None = None
    note: str | None = None


def _ranking_note(subject: str, shown: int, total: int, where: str) -> str | None:
    """Dernière ligne d'un onglet de classement : sur combien il porte, et où est le reste."""
    if total <= shown:
        return None
    return (
        f"Classement borné : {_grouped(shown)} {subject} sur {_grouped(total)}. "
        f"La totalité est dans {where}."
    )


def _sheet_scope(report: ReportData) -> _Sheet | None:
    scope = report.scope
    if not scope.incomplete:
        return None
    return _Sheet(
        SCOPE_SHEET,
        ["Inventaire incomplet — à lire avant toute décision de suppression"],
        [
            [scope.headline()],
            *[[f"Non parcouru : {cible}"] for cible in scope.skipped_targets],
            *[[message] for message in scope.warnings],
        ],
    )


def _sheet_synthese(report: ReportData) -> _Sheet:
    o, scope = report.overview, report.scope
    return _Sheet(
        "Synthèse",
        ["Indicateur", "Valeur", "Détail"],
        [
            *(
                [["Périmètre", "INCOMPLET", scope.headline()]]
                if scope.incomplete
                else [["Périmètre", "complet", "toutes les cibles demandées ont été parcourues"]]
            ),
            ["Base", o.db_path, ""],
            ["Généré le", o.generated_at, ""],
            ["Modèle", o.model, ""],
            ["Prompt", o.prompt_name, o.prompt_hash],
            ["Fichiers inventoriés", o.total_files, views.format_bytes(o.total_bytes)],
            ["Volume total (octets)", o.total_bytes, ""],
            ["Analysés", o.analyzed, f"{views.percent(o.analyzed, o.total_files)} %"],
            ["À analyser", o.pending, ""],
            ["Exclus", o.excluded, ""],
            ["En erreur", o.errors, ""],
            ["Familles de doublons", o.duplicate_families, ""],
            [
                "Octets récupérables (doublons)",
                o.duplicate_reclaimable_bytes,
                views.format_bytes(o.duplicate_reclaimable_bytes),
            ],
            [f"Fichiers non accédés depuis {o.stale_years} ans", o.stale_files, ""],
            ["Octets non accédés", o.stale_bytes, views.format_bytes(o.stale_bytes)],
            ["Fichiers sensibles (C2/C3)", o.sensitive_files, ""],
            ["RGPD élevé ou critique", o.rgpd_at_risk, ""],
            ["Fichiers à conserver", o.retention_files, ""],
            [
                "dont durée non déterminée",
                report.retention.undetermined_files,
                "à trancher par un humain : jamais « échu »",
            ],
            ["dont conservation échue", report.retention.expired_files, ""],
            ["Candidats au nettoyage", o.cleanup_files, views.format_bytes(o.cleanup_bytes)],
            ["Octets libérables", o.cleanup_bytes, ""],
            ["Vérifiés par un humain", o.reviewed, ""],
        ],
    )


def _sheet_doublons(report: ReportData) -> _Sheet:
    families = report.duplicates.families
    return _Sheet(
        "Doublons",
        ["Famille", "Empreinte", "Taille unitaire (o)", "Copies", "Octets récupérables", "Chemins"],
        (
            [
                f.family_id,
                f.fast_hash,
                f.size_bytes,
                f.copies,
                f.reclaimable_bytes,
                " | ".join(f.paths),
            ]
            for f in families
        ),
        formats={3: "#,##0", 5: "#,##0"},
        note=" ".join(
            filter(
                None,
                (
                    views.DUPLICATE_CAUTION,
                    _ranking_note(
                        "familles affichées",
                        len(families),
                        report.totals.get("duplicates", len(families)),
                        "« export --format powerbi » (duplicates.csv, un exemplaire par ligne)",
                    ),
                ),
            )
        ),
    )


def _sheet_anciennete(report: ReportData) -> _Sheet:
    return _Sheet(
        "Ancienneté",
        [
            "Seuil (ans)",
            "Antérieur au",
            "Fichiers non accédés",
            "Octets non accédés",
            "Fichiers non modifiés",
            "Octets non modifiés",
        ],
        (
            [
                b.years,
                b.cutoff,
                b.not_accessed_files,
                b.not_accessed_bytes,
                b.not_modified_files,
                b.not_modified_bytes,
            ]
            for b in report.stale
        ),
        formats={2: "DD/MM/YYYY", 4: "#,##0", 6: "#,##0"},
    )


def _sheet_sensibles(report: ReportData) -> _Sheet:
    return _Sheet(
        "Sensibles",
        [
            "file_id",
            "Chemin",
            "Propriétaire",
            "Taille (o)",
            "Sécurité",
            "Confiance sécurité",
            "RGPD",
            "Confiance RGPD",
            "Résumé",
            "Justification",
            "Revue",
        ],
        (
            [
                f.file_id,
                f.path,
                f.owner,
                f.size_bytes,
                f.security,
                f.security_confidence,
                f.rgpd,
                f.rgpd_confidence,
                f.resume,
                f.justification,
                f.review_status,
            ]
            for f in report.sensitive
        ),
        formats={4: "#,##0"},
        note=_ranking_note(
            "fichiers affichés",
            len(report.sensitive),
            report.totals.get("sensitive", len(report.sensitive)),
            "l'onglet « Fichiers » de ce classeur et dans « export --format powerbi »",
        ),
    )


def _sheet_conservation(report: ReportData) -> _Sheet:
    return _Sheet(
        "Conservation",
        [
            "file_id",
            "Chemin",
            "Propriétaire",
            "Taille (o)",
            "Durée (ans)",
            "Fondement",
            "Dernière écriture",
            "Fin de conservation",
            "Échu",
            "Justification",
        ],
        (
            [
                r.file_id,
                r.path,
                r.owner,
                r.size_bytes,
                views.RETENTION_UNDETERMINED if r.undetermined else r.years,
                views.RETENTION_BASIS_LABELS.get(r.basis, r.basis),
                r.last_write_time,
                _date_cell(r.end_date),
                "oui" if r.expired else "non",
                r.justification,
            ]
            for r in report.retention.rows
        ),
        formats={4: "#,##0", 8: "DD/MM/YYYY"},
    )


def _sheet_nettoyage(report: ReportData) -> _Sheet:
    return _Sheet(
        "Nettoyage",
        ["file_id", "Chemin", "Propriétaire", "Taille (o)", "Dernier accès", "Sécurité"],
        (
            [r.file_id, r.path, r.owner, r.size_bytes, r.access_time, r.security]
            for r in report.cleanup.rows
        ),
        formats={4: "#,##0"},
    )


def _sheet_revues(report: ReportData) -> _Sheet:
    reviews = report.reviews
    return _Sheet(
        "Revues",
        ["Indicateur", "Valeur"],
        [
            ["À vérifier", reviews.to_review],
            ["Validés", reviews.validated],
            ["Corrigés", reviews.corrected],
            ["Non revus", reviews.not_reviewed],
            ["Analysés", reviews.analyzed],
            ["Avancement (%)", reviews.percent_reviewed],
        ]
        + [
            [f"Écart — {d.path}", f"{d.llm_security}→{d.corrected_security or '—'}"]
            for d in reviews.discrepancies
        ],
    )


def _sheet_erreurs(report: ReportData) -> _Sheet:
    status = report.status
    return _Sheet(
        "Erreurs",
        ["Raison", "Fichiers", "Octets"],
        [[g.label, g.files, g.bytes] for g in status.reasons]
        + [[f"statut : {k}", v, status.bytes.get(k, 0)] for k, v in status.counts.items()],
        formats={3: "#,##0"},
        note=(
            f"Motifs bornés : {_grouped(len(status.reasons))} sur "
            f"{_grouped(status.reasons_total)} affichés. Ce sont des raisons de "
            "non-analyse : les taire ferait passer pour propre un partage mal couvert."
            if status.reasons_hidden
            else None
        ),
    )


def write_workbook(
    db: Database,
    path: Path,
    *,
    today: date | None = None,
    data: ReportData | None = None,
    on_warning: Callable[[str], None] | None = None,
    max_sheet_rows: int = MAX_SHEET_ROWS,
) -> Path:
    """Écrit le classeur `path` et le rend.

    Le classeur est en écriture seule : l'onglet « Fichiers » est écrit ligne à
    ligne depuis la base, sans liste intermédiaire. Les largeurs de colonnes sont
    déduites des `WIDTH_SAMPLE_ROWS` premières lignes de chaque onglet.

    Un onglet qui dépasserait `max_sheet_rows` (limite d'Excel, 1 048 576 lignes)
    est tronqué : sa dernière ligne porte un avertissement en clair, et le même
    message part vers `on_warning` (par défaut le journal, affiché par la CLI).
    Les formats `powerbi` et `csv` restent la sortie complète.

    **« Conservation » et « Nettoyage » portent la totalité des lignes** — c'est
    le sens de `collect(actions=None)`. Chacune de ces lignes est une décision de
    suppression : annoncer « 182 346 fichiers, 7,5 To libérables » puis n'en
    donner que 50, dans le format même vers lequel le rapport renvoie, rendait le
    gain inatteignable. « Doublons » et « Sensibles » restent des classements
    bornés, et leur dernière ligne dit désormais sur combien ils portent.

    Un onglet = une fonction `_sheet_*` (pure) ; cette fonction ne fait que les
    écrire dans l'ordre.
    """
    report = data if data is not None else collect(db, today=today, actions=None)
    warn = on_warning if on_warning is not None else logger.warning
    warnings: list[str] = []
    workbook = Workbook(write_only=True)
    file_headers, file_rows = _file_rows(db)
    sheets: list[_Sheet | None] = [
        _sheet_scope(report),
        _sheet_synthese(report),
        _Sheet("Fichiers", file_headers, file_rows),
        _sheet_doublons(report),
        _sheet_anciennete(report),
        _sheet_sensibles(report),
        _sheet_conservation(report),
        _sheet_nettoyage(report),
        _sheet_revues(report),
        _sheet_erreurs(report),
    ]
    for sheet in sheets:
        if sheet is None:
            continue
        message = _add_sheet(
            workbook,
            sheet.title,
            sheet.headers,
            sheet.rows,
            formats=sheet.formats,
            max_rows=max_sheet_rows,
            note=sheet.note,
        )
        if message is not None:
            warnings.append(message)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(path))
    for message in warnings:
        warn(message)
    return path


__all__ = ["MAX_SHEET_ROWS", "SHEETS", "WIDTH_SAMPLE_ROWS", "write_workbook"]
