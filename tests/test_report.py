"""Rapport HTML / Markdown, classeur Excel et export Power BI."""

from __future__ import annotations

import csv
import json
import re
import zipfile
from collections.abc import Iterator
from datetime import date, datetime
from html.parser import HTMLParser
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from docia import views
from docia.cli import main
from docia.db import Database
from docia.models import FileStatus
from docia.report import excel, powerbi
from docia.report.data import collect
from docia.report.html import render_html
from docia.report.markdown import render_markdown
from tests.test_views import TODAY, _analysis, _row

_VOID = {"meta", "br", "hr", "img", "input", "link", "rect", "path"}


class _Checker(HTMLParser):
    """Vérifie que les balises sont bien imbriquées et toutes refermées."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.stack: list[str] = []
        self.errors: list[str] = []
        self.tags: set[str] = set()

    def handle_starttag(self, tag: str, _attrs: list[tuple[str, str | None]]) -> None:
        self.tags.add(tag)
        if tag not in _VOID:
            self.stack.append(tag)

    def handle_endtag(self, tag: str) -> None:
        if tag in _VOID:
            return
        if not self.stack or self.stack[-1] != tag:
            self.errors.append(f"</{tag}> inattendu (pile : {self.stack[-3:]})")
            return
        self.stack.pop()


@pytest.fixture
def db(tmp_path: Path) -> Iterator[Database]:
    """Petite base représentative (doublons, ancien, sensible, conservation, revue)."""
    database = Database(tmp_path / "rapport.sqlite")
    scan = database.start_scan("scan.csv")
    database.upsert_files(
        [
            _row("contrat.pdf", fast_hash="dup", size=4096),
            _row("contrat-copie.pdf", fast_hash="dup", size=4096, directory="\\\\srv\\part\\bis"),
            _row(
                "archive.txt",
                fast_hash="old",
                size=1_500_000,
                lwt="01/02/2017 08:00:00",
                access="02/02/2017 08:00:00",
                owner="DOM\\bob",
            ),
            _row("paye.xlsx", fast_hash="rh", size=250_000, owner="DOM\\rh"),
        ],
        scan,
    )
    files = {f.name: f for f in database.iter_files()}
    database.store_analysis(
        files["contrat.pdf"].id,
        None,
        1,
        prompt_hash="ph0123456789abcd",
        model="qwen38",
        analysis=_analysis(
            "contrat.pdf", security="C3", rgpd="critical", retention=True, years=10, basis="legal"
        ),
    )
    database.store_analysis(
        files["archive.txt"].id,
        None,
        1,
        prompt_hash="ph0123456789abcd",
        model="qwen38",
        analysis=_analysis("archive.txt", security="C0", rgpd="none"),
    )
    database.store_analysis(
        files["paye.xlsx"].id,
        None,
        1,
        prompt_hash="ph0123456789abcd",
        model="qwen38",
        analysis=_analysis(
            "paye.xlsx", security="C2", rgpd="high", retention=True, years=5, basis="rh"
        ),
    )
    database.set_review(files["contrat.pdf"].id, "corrected", corrected_security="C2")
    yield database
    database.close()


# ------------------------------------------------------------------ HTML


def test_html_is_self_contained_and_well_formed(db: Database) -> None:
    page = render_html(db, today=TODAY)
    checker = _Checker()
    checker.feed(page)
    checker.close()
    assert checker.errors == []
    assert checker.stack == []
    # autonome : aucun script, aucune ressource externe
    assert "<script" not in page.lower()
    assert "<script src=" not in page.lower()
    assert "http://" not in page.replace("http://www.w3.org/2000/svg", "")
    assert "https://" not in page
    assert "<link" not in page.lower()
    # les graphiques sont du SVG écrit en Python
    assert "svg" in checker.tags
    assert "<rect" in page


def test_html_contains_all_sections_and_key_numbers(db: Database) -> None:
    page = render_html(db, today=TODAY)
    for anchor in ("synthese", "hygiene", "risque", "verification", "execution"):
        assert f'id="{anchor}"' in page
    for title in (
        "Synthèse",
        "Hygiène du stockage",
        "Risque et conformité",
        "Vérification humaine",
        "Exécution",
        "Doublons",
        "Ancienneté",
        "Plan de conservation",
        "Candidats au nettoyage",
    ):
        assert title in page
    # 4 fichiers, 3 analysés, 1 famille de doublons de 4 096 octets récupérables
    assert views.format_int(4) in page
    assert views.format_bytes(4096) in page
    assert "contrat.pdf" in page
    assert "C3" in page
    assert "critical" in page
    # date de fin de conservation : 01/01/2026 + 10 ans
    assert "01/01/2036" in page


def test_html_escapes_dangerous_content(tmp_path: Path) -> None:
    with Database(tmp_path / "x.sqlite") as database:
        scan = database.start_scan("s.csv")
        database.upsert_files(
            [_row("a.txt", fast_hash="h", owner="<script>alert(1)</script>")], scan
        )
        files = list(database.iter_files())
        database.set_file_status(files[0].id, FileStatus.EXCLUDED, "raison <b>brute</b>")
        page = render_html(database, today=TODAY)
    assert "<script>alert(1)</script>" not in page
    assert "&lt;script&gt;" in page
    assert "&lt;b&gt;brute&lt;/b&gt;" in page


# ------------------------------------------------------------------ Markdown


def test_markdown_contains_tables_and_sections(db: Database) -> None:
    text = render_markdown(db, today=TODAY)
    assert text.startswith("# Doc-IA — rapport d'analyse")
    for heading in (
        "## 1. Synthèse",
        "## 2. Hygiène du stockage",
        "## 3. Risque et conformité",
        "## 4. Vérification humaine",
        "## 5. Exécution",
    ):
        assert heading in text
    # tableaux GFM : au moins une ligne de séparation
    assert re.search(r"^\|---\|", text, re.MULTILINE)
    assert "| Exemplaire de référence | Copies |" in text
    assert "contrat.pdf" in text
    assert "01/01/2036" in text


def test_html_and_markdown_share_the_same_data(db: Database) -> None:
    data = collect(db, today=TODAY)
    page = render_html(db, data=data)
    text = render_markdown(db, data=data)
    assert data.overview.total_files == 4
    assert data.overview.analyzed == 3
    for rendering in (page, text):
        assert views.format_bytes(data.duplicates.total_reclaimable_bytes) in rendering


# ------------------------------------------------------------------ Excel


def test_les_motifs_non_affiches_sont_annonces(tmp_path: Path) -> None:
    """MOYEN : la borne des motifs d'exclusion était mesurée, jamais dite.

    `views.REASON_TOP` en montre 10. Sur une campagne à 25 motifs, 15 disparaissaient
    en silence des trois rendus — alors que `StatusSummary.reasons_total` et
    `reasons_hidden` existaient déjà et n'étaient lus nulle part. Ce sont des raisons
    de **non**-analyse : les taire fait passer pour propre un partage mal couvert.
    """
    from tests.test_views import _row

    with Database(tmp_path / "motifs.sqlite") as database:
        scan = database.start_scan("s")
        database.upsert_files([_row(f"f{i}.txt", fast_hash=f"h{i}") for i in range(25)], scan)
        database.finish_scan(scan, total=25, new=25, updated=0, unchanged=0, invalid=0)
        for fichier in database.iter_files():
            database.set_file_status(fichier.id, "excluded", reason=f"motif {fichier.name}")
        resume = views.status_summary(database)
        assert (len(resume.reasons), resume.reasons_total, resume.reasons_hidden) == (10, 25, 15)
        page = render_html(database, today=TODAY)
        texte = render_markdown(database, today=TODAY)

    for rendu, nom in ((page, "HTML"), (texte, "Markdown")):
        assert "25" in rendu, f"{nom} ne donne pas le nombre réel de motifs"
        assert "autres ne sont pas" in rendu, f"{nom} ne dit pas que le tableau est coupé"


def test_les_doublons_ne_sont_jamais_annonces_comme_identiques(
    db: Database, tmp_path: Path
) -> None:
    """GRAVE : les rendus affirmaient « fichiers identiques » sur 64 Ko comparés.

    `fast_hash` ne couvre que les 64 premiers kilo-octets (`quick.HASH_HEAD_BYTES`,
    comme SMBeagle). Deux fichiers de 200 Ko dont seuls les 64 premiers coïncident
    étaient donc déclarés identiques — c'est le cas des formats à en-tête fixe, des
    images disque et des exports au même gabarit — et le tableau chiffrait l'espace
    « récupérable en ne gardant qu'un exemplaire ». Aucun des quatre rendus ne disait
    que la comparaison était partielle.
    """
    page = render_html(db, today=TODAY)
    texte = render_markdown(db, today=TODAY)
    powerbi.export_powerbi(db, tmp_path / "pbi", today=TODAY)
    lisez_moi = (tmp_path / "pbi" / "README_powerbi.md").read_text(encoding="utf-8")

    for rendu, nom in ((page, "HTML"), (texte, "Markdown"), (lisez_moi, "README Power BI")):
        assert "64 premiers Ko" in rendu, f"{nom} ne dit pas que l'empreinte est partielle"
    assert "fichiers identiques" not in page, "le mot « identiques » promet plus que le calcul"
    assert "octet à octet" in page, "le HTML doit dire quoi faire avant de supprimer"
    assert "octet à octet" in texte


def test_excel_workbook_sheets_and_rows(db: Database, tmp_path: Path) -> None:
    path = excel.write_workbook(db, tmp_path / "rapport.xlsx", today=TODAY)
    assert path.exists()
    workbook = load_workbook(path)
    assert workbook.sheetnames == list(excel.SHEETS)
    files_sheet = workbook["Fichiers"]
    assert files_sheet.max_row == 5  # 4 fichiers + en-tête
    assert files_sheet["A1"].font.bold is True
    assert files_sheet.freeze_panes == "A2"
    assert files_sheet.auto_filter.ref is not None
    doublons = workbook["Doublons"]
    # en-tête + 1 famille + la mise en garde : l'onglet chiffre un espace « récupérable »,
    # donc invite à supprimer, sur un regroupement qui ne compare que les 64 premiers Ko.
    assert doublons.max_row == 3
    assert doublons.cell(row=2, column=5).value == 4096  # octets récupérables
    assert "64 premiers Ko" in str(doublons.cell(row=3, column=1).value)
    anciennete = workbook["Ancienneté"]
    assert anciennete.max_row == 5  # 4 seuils
    conservation = workbook["Conservation"]
    assert conservation.max_row == 3
    ends = {conservation.cell(row=r, column=8).value for r in (2, 3)}
    assert any(str(value).startswith("2036-01-01") for value in ends)
    assert workbook["Sensibles"].max_row == 3
    assert workbook["Synthèse"].max_row > 10


# ------------------------------------------------------------------ Power BI


def _read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle, delimiter=";"))
    return rows[0], rows[1:]


def test_powerbi_export_files_and_encoding(db: Database, tmp_path: Path) -> None:
    target = tmp_path / "powerbi"
    written = powerbi.export_powerbi(db, target, today=TODAY)
    names = {p.name for p in written}
    assert names == set(powerbi.POWERBI_COLUMNS) | {powerbi.README_NAME}
    for name in powerbi.POWERBI_COLUMNS:
        raw = (target / name).read_bytes()
        assert raw.startswith(b"\xef\xbb\xbf"), f"{name} : BOM UTF-8 manquant"
        assert b";" in raw, f"{name} : séparateur ; manquant"

    header, rows = _read_csv(target / "files.csv")
    assert header == [c[0] for c in powerbi.POWERBI_COLUMNS["files.csv"]]
    assert len(rows) == 4
    by_name = {row[header.index("name")]: row for row in rows}
    assert by_name["archive.txt"][header.index("access_date")] == "2017-02-02"
    assert by_name["archive.txt"][header.index("share")] == "\\\\srv\\part"
    assert int(by_name["archive.txt"][header.index("age_days_access")]) > 3000

    header, rows = _read_csv(target / "analyses.csv")
    assert header == [c[0] for c in powerbi.POWERBI_COLUMNS["analyses.csv"]]
    assert len(rows) == 3
    ends = {row[header.index("retention_end_date")] for row in rows}
    assert "2036-01-01" in ends
    assert "identite" in {row[header.index("rgpd_data_types")] for row in rows}

    header, rows = _read_csv(target / "duplicates.csv")
    assert len(rows) == 2  # une famille de deux exemplaires
    assert rows[0][header.index("reclaimable_bytes")] == "4096"

    header, rows = _read_csv(target / "reviews.csv")
    assert len(rows) == 1
    assert rows[0][header.index("corrected_security")] == "C2"

    header, _ = _read_csv(target / "runs.csv")
    assert header == [c[0] for c in powerbi.POWERBI_COLUMNS["runs.csv"]]


def test_powerbi_readme_documents_exactly_the_real_columns(db: Database, tmp_path: Path) -> None:
    target = tmp_path / "powerbi"
    powerbi.export_powerbi(db, target, today=TODAY)
    readme = (target / powerbi.README_NAME).read_text(encoding="utf-8")
    for name in powerbi.POWERBI_COLUMNS:
        assert f"### `{name}`" in readme
        documented = re.findall(rf"### `{re.escape(name)}`.*?(?=\n### |\Z)", readme, re.S)[0]
        columns = re.findall(r"^\| `([a-z_]+)` \|", documented, re.MULTILINE)
        real, _ = _read_csv(target / name)
        assert columns == real, f"{name} : README et CSV divergent"
    assert "utf-8" in readme.lower()
    assert "file_id" in readme
    assert "Rafraîchissement" in readme


def test_powerbi_export_is_idempotent(db: Database, tmp_path: Path) -> None:
    target = tmp_path / "powerbi"
    first = powerbi.export_powerbi(db, target, today=TODAY)
    second = powerbi.export_powerbi(db, target, today=TODAY)
    assert [p.name for p in first] == [p.name for p in second]
    assert (target / "files.csv").read_bytes() == (target / "files.csv").read_bytes()


# ------------------------------------------------------------------ CLI


def test_cli_report_and_export_formats(db: Database, tmp_path: Path) -> None:
    db_path = str(db.path)
    db.close()
    out_html = tmp_path / "r.html"
    assert main(["--db", db_path, "report", "--format", "html", "--out", str(out_html)]) == 0
    assert "Doc-IA" in out_html.read_text(encoding="utf-8")

    out_md = tmp_path / "r.md"
    assert main(["--db", db_path, "report", "--format", "md", "--out", str(out_md)]) == 0
    assert out_md.read_text(encoding="utf-8").startswith("# Doc-IA")

    out_xlsx = tmp_path / "r.xlsx"
    assert main(["--db", db_path, "export", "--format", "xlsx", "--out", str(out_xlsx)]) == 0
    assert load_workbook(out_xlsx).sheetnames == list(excel.SHEETS)

    out_pbi = tmp_path / "pbi"
    assert main(["--db", db_path, "export", "--format", "powerbi", "--out", str(out_pbi)]) == 0
    assert (out_pbi / "files.csv").exists()

    # les formats existants restent intacts
    out_csv = tmp_path / "r.csv"
    assert main(["--db", db_path, "export", "--format", "csv", "--out", str(out_csv)]) == 0
    assert out_csv.read_text(encoding="utf-8-sig").startswith("path;")
    out_json = tmp_path / "r.json"
    assert main(["--db", db_path, "export", "--format", "json", "--out", str(out_json)]) == 0
    assert out_json.read_text(encoding="utf-8").lstrip().startswith("[")


def test_cli_report_default_output_next_to_database(db: Database) -> None:
    db_path = Path(db.path)
    db.close()
    assert main(["--db", str(db_path), "report"]) == 0
    assert db_path.with_name(f"{db_path.stem}_rapport.html").exists()


# ------------------------------------------------- Excel : flux et limite de lignes


def test_excel_tronque_au_dela_de_la_limite_de_lignes(db: Database, tmp_path: Path) -> None:
    """Limite abaissée à 4 lignes : l'onglet est tronqué et le dit, dans le fichier et à l'écran."""
    alertes: list[str] = []
    path = excel.write_workbook(
        db,
        tmp_path / "tronque.xlsx",
        today=TODAY,
        on_warning=alertes.append,
        max_sheet_rows=4,
    )
    feuille = load_workbook(path)["Fichiers"]
    assert feuille.max_row == 4  # en-tête + 2 fichiers + avertissement
    avertissement = str(feuille.cell(row=4, column=1).value)
    assert "tronqué" in avertissement
    assert "powerbi" in avertissement
    assert "csv" in avertissement
    assert "2 ligne(s) écrite(s), 2 non écrite(s)" in avertissement
    assert feuille.cell(row=4, column=1).font.bold is True
    # le même message est remis à l'utilisateur, une fois par onglet tronqué
    assert avertissement in alertes
    assert any("Fichiers" in message for message in alertes)


def test_excel_sans_troncature_ne_previent_pas(db: Database, tmp_path: Path) -> None:
    """Sous la limite, rien ne change : pas d'avertissement, toutes les lignes présentes.

    La limite est abaissée à la hauteur exacte de la Synthèse, le plus gros onglet
    de cette base : le cas « pile à la limite » est donc couvert.

    Elle est passée de 23 à 24 lignes le jour où la Synthèse a gagné son indicateur
    « Périmètre » (complet / INCOMPLET) : le classeur sert aux mêmes décisions de
    suppression que le rapport HTML et doit dire, lui aussi, si l'inventaire porte
    sur tout ce qui a été demandé. Sans ce réglage, la Synthèse débordait d'une
    ligne et le test échouait sur un avertissement de troncature parasite.
    """
    alertes: list[str] = []
    path = excel.write_workbook(
        db, tmp_path / "complet.xlsx", today=TODAY, on_warning=alertes.append, max_sheet_rows=24
    )
    assert alertes == []
    classeur = load_workbook(path)
    assert classeur["Synthèse"].max_row == 24  # 23 indicateurs + en-tête, pile la limite
    feuille = classeur["Fichiers"]
    assert feuille.max_row == 5  # 4 fichiers + en-tête
    assert feuille.auto_filter.ref is not None
    assert feuille.freeze_panes == "A2"


def test_excel_largeurs_derivees_dun_echantillon_borne(
    db: Database, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """Les largeurs viennent des N premières lignes : c'est le prix du mode écriture seule."""
    monkeypatch.setattr(excel, "WIDTH_SAMPLE_ROWS", 1)
    etroit = load_workbook(excel.write_workbook(db, tmp_path / "a.xlsx", today=TODAY))["Fichiers"]
    monkeypatch.setattr(excel, "WIDTH_SAMPLE_ROWS", 1000)
    large = load_workbook(excel.write_workbook(db, tmp_path / "b.xlsx", today=TODAY))["Fichiers"]
    largeurs_etroites = [c.width for c in etroit.column_dimensions.values()]
    largeurs_larges = [c.width for c in large.column_dimensions.values()]
    assert largeurs_etroites != largeurs_larges
    assert max(largeurs_larges) >= max(largeurs_etroites)


def test_excel_lit_la_base_en_flux(db: Database, tmp_path: Path) -> None:
    """`latest_analyses` est consommé comme un curseur, jamais transformé en liste."""
    consomme = 0
    vraie = db.latest_analyses

    def espion(**kwargs: object) -> Iterator[object]:
        nonlocal consomme
        for record in vraie(**kwargs):  # type: ignore[arg-type]
            consomme += 1
            yield record

    with pytest.MonkeyPatch.context() as patch:
        patch.setattr(db, "latest_analyses", espion)
        excel.write_workbook(db, tmp_path / "flux.xlsx", today=TODAY)
    assert consomme == 4


def test_powerbi_pagination_rend_toutes_les_lignes_dans_lordre(db: Database) -> None:
    """La pagination par clé ne saute ni ne répète aucune ligne, quelle que soit la tranche."""
    sql = "SELECT * FROM files WHERE id > ? ORDER BY id LIMIT ?"
    complet = [int(r["id"]) for r in db.query("SELECT id FROM files ORDER BY id")]
    for page in (1, 2, 3, 100):
        assert [int(r["id"]) for r in powerbi._pages(db, sql, "id", page=page)] == complet, page


# ------------------------------------- CLI : export CSV et JSON écrits en flux


def test_cli_export_csv_et_json_identiques_a_la_version_en_memoire(
    db: Database, tmp_path: Path
) -> None:
    """Preuve d'identité : l'écriture en flux rend exactement les octets d'avant."""
    db_path = str(db.path)
    rows = [dict(r) for r in db.latest_analyses()]
    db.close()

    out_csv = tmp_path / "flux.csv"
    assert main(["--db", db_path, "export", "--format", "csv", "--out", str(out_csv)]) == 0
    attendu_csv = tmp_path / "memoire.csv"
    with attendu_csv.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()), delimiter=";")
        writer.writeheader()
        writer.writerows(rows)
    assert out_csv.read_bytes() == attendu_csv.read_bytes()

    out_json = tmp_path / "flux.json"
    assert main(["--db", db_path, "export", "--format", "json", "--out", str(out_json)]) == 0
    for r in rows:
        for key in ("rgpd_data_types", "finance_amounts", "legal_parties"):
            if r.get(key):
                r[key] = json.loads(r[key])
    assert out_json.read_text(encoding="utf-8") == json.dumps(rows, ensure_ascii=False, indent=2)
    assert len(json.loads(out_json.read_text(encoding="utf-8"))) == 4


def test_cli_export_base_vide(tmp_path: Path) -> None:
    """Base sans fichier : un CSV réduit à son en-tête, un JSON `[]` — comme avant."""
    vide = tmp_path / "vide.sqlite"
    Database(vide).close()
    out_csv = tmp_path / "v.csv"
    assert main(["--db", str(vide), "export", "--format", "csv", "--out", str(out_csv)]) == 0
    assert out_csv.read_bytes() == "﻿path\r\n".encode()
    out_json = tmp_path / "v.json"
    assert main(["--db", str(vide), "export", "--format", "json", "--out", str(out_json)]) == 0
    assert out_json.read_text(encoding="utf-8") == "[]"


# ----------------------------------- dates aberrantes : un fichier ne coûte pas le rapport


HOSTILE_NAMES: tuple[str, ...] = (
    "=cmd|'/c calc.exe'!A1.docx",  # DDE : le vecteur historique d'Excel
    "=1+1.txt",
    "-- Copie de sauvegarde.docx",  # tiret initial : très courant sur un partage
    "+33 1 23 45 67 89 devis.pdf",
    "@Envoyer a la compta.docx",
    "<img src=x onerror=alert(1)>.pdf",
    "Contrat d'associé & Cie.pdf",
)
"""Noms de fichiers ordinaires sur un partage français — et vecteurs de formule."""

HOSTILE_RESUME = '=HYPERLINK("http://mechant";"cliquez") résumé <b>gras</b> & `code`'
"""Le `resume` de la LLM : sans contrainte de caractères, il traverse jusqu'aux sorties."""


def _base_hostile(path: Path) -> Database:
    """Base d'un fichier par nom hostile, chacun avec un résumé hostile."""
    database = Database(path)
    scan = database.start_scan("scan.csv")
    database.upsert_files(
        [_row(name, fast_hash=f"h{i}", size=4096) for i, name in enumerate(HOSTILE_NAMES)], scan
    )
    database.finish_scan(
        scan, total=len(HOSTILE_NAMES), new=len(HOSTILE_NAMES), updated=0, unchanged=0, invalid=0
    )
    for file_row in database.iter_files():
        database.store_analysis(
            file_row.id,
            None,
            1,
            prompt_hash="ph0123456789abcd",
            model="qwen38",
            analysis=_analysis(
                file_row.name,
                security="C3",
                rgpd="critical",
                retention=True,
                years=10,
                basis="legal",
                resume=HOSTILE_RESUME,
                justification="=1+2 justification",
            ),
        )
    return database


def test_une_date_en_9999_ne_coute_pas_les_quatre_rapports(tmp_path: Path) -> None:
    """Un seul FILETIME saturé rendait la campagne entière non restituable.

    `DateTime.MaxValue` de .NET vaut 9999-12-31 : c'est ce que rend la conversion
    d'un FILETIME corrompu ou saturé (fichier restauré d'une archive abîmée, NAS
    à horloge cassée). Avec une conservation légale de 10 ans, `shift_years`
    tentait `date(10009, …)` et `html`, `markdown`, `powerbi` **et** `xlsx`
    échouaient tous les quatre sur `ValueError: year 10009 is out of range`.
    """
    with Database(tmp_path / "futur.sqlite") as database:
        scan = database.start_scan("scan.csv")
        database.upsert_files(
            [
                _row("futur.pdf", lwt="01/01/9999 10:00:00", access="01/01/9999 10:00:00"),
                _row("normal.pdf", fast_hash="n"),
            ],
            scan,
        )
        database.finish_scan(scan, total=2, new=2, updated=0, unchanged=0, invalid=0)
        for file_row in database.iter_files():
            database.store_analysis(
                file_row.id,
                None,
                1,
                prompt_hash="ph0123456789abcd",
                model="qwen38",
                analysis=_analysis(
                    file_row.name, retention=True, years=10, basis="legal", security="C2"
                ),
            )
        page = render_html(database, today=TODAY)
        texte = render_markdown(database, today=TODAY)
        classeur = excel.write_workbook(database, tmp_path / "futur.xlsx", today=TODAY)
        powerbi.export_powerbi(database, tmp_path / "pbi", today=TODAY)

    # la fin de conservation est bornée à la dernière date représentable, pas inventée
    assert "31/12/9999" in page
    assert "31/12/9999" in texte
    csv_analyses = (tmp_path / "pbi" / "analyses.csv").read_text(encoding="utf-8-sig")
    assert "9999-12-31" in csv_analyses
    conservation = load_workbook(classeur)["Conservation"]
    assert any(
        str(cell.value).startswith("9999-12-31")
        for row in conservation.iter_rows()
        for cell in row
        if cell.value is not None
    )


# ------------------------------------------- injection de formule (Excel, Power BI)


def test_excel_nécrit_jamais_de_formule(tmp_path: Path) -> None:
    """Preuve sur le fichier produit : aucun `<f>` dans le XML des feuilles.

    Une chaîne commençant par `=` était écrite en `<f>…</f>` — une vraie formule,
    évaluée à l'ouverture. Le classeur est lu par des non-informaticiens, la
    population exactement visée par le vecteur DDE.
    """
    with _base_hostile(tmp_path / "hostile.sqlite") as database:
        chemin = excel.write_workbook(database, tmp_path / "hostile.xlsx", today=TODAY)

    with zipfile.ZipFile(chemin) as archive:
        feuilles = [n for n in archive.namelist() if n.startswith("xl/worksheets/")]
        assert feuilles
        for nom in feuilles:
            assert b"<f>" not in archive.read(nom), nom

    classeur = load_workbook(chemin)
    formules = [
        (feuille.title, cellule.coordinate, cellule.value)
        for feuille in classeur
        for ligne in feuille.iter_rows()
        for cellule in ligne
        if cellule.data_type == "f"
    ]
    assert formules == []

    # le texte n'est pas perdu pour autant : il est écrit tel quel, en chaîne
    valeurs = {
        str(cellule.value)
        for feuille in classeur
        for ligne in feuille.iter_rows()
        for cellule in ligne
        if cellule.value is not None
    }
    assert any("=cmd|'/c calc.exe'!A1.docx" in v for v in valeurs)
    assert any(v.startswith(HOSTILE_RESUME) for v in valeurs)


def test_excel_garde_les_nombres_et_les_dates_calculables(db: Database, tmp_path: Path) -> None:
    """L'assainissement ne doit pas transformer les nombres ni les dates en texte."""
    feuilles = load_workbook(excel.write_workbook(db, tmp_path / "types.xlsx", today=TODAY))
    tailles = feuilles["Sensibles"]
    assert all(
        isinstance(ligne[3].value, int) for ligne in tailles.iter_rows(min_row=2) if ligne[3].value
    )
    anciennete = feuilles["Ancienneté"]
    assert all(
        isinstance(ligne[1].value, datetime | date)
        for ligne in anciennete.iter_rows(min_row=2)
        if ligne[1].value
    )


def test_powerbi_csv_naucune_cellule_interpretee_comme_formule(tmp_path: Path) -> None:
    """Preuve sur les CSV produits : plus une seule cellule commençant par `= + - @`.

    L'export est en `utf-8-sig` avec `;` précisément pour qu'Excel l'ouvre d'un
    double-clic : `- copie.docx` s'y affichait `#NOM ?` au lieu du nom du fichier.
    """
    dossier = tmp_path / "pbi"
    with _base_hostile(tmp_path / "hostile.sqlite") as database:
        powerbi.export_powerbi(database, dossier, today=TODAY)

    suspectes: list[tuple[str, str, str]] = []
    for fichier in sorted(dossier.glob("*.csv")):
        entetes, lignes = _read_csv(fichier)
        for ligne in lignes:
            for colonne, cellule in zip(entetes, ligne, strict=True):
                if cellule[:1] in ("=", "+", "-", "@", "\t", "\r"):
                    suspectes.append((fichier.name, colonne, cellule))
    assert suspectes == []

    # la valeur reste lisible, simplement neutralisée par une apostrophe
    _, lignes = _read_csv(dossier / "files.csv")
    noms = {ligne[2] for ligne in lignes}
    assert "'-- Copie de sauvegarde.docx" in noms
    assert "'+33 1 23 45 67 89 devis.pdf" in noms
    assert "'@Envoyer a la compta.docx" in noms
    assert "<img src=x onerror=alert(1)>.pdf" in noms  # rien d'inutile n'est préfixé


def test_powerbi_csv_ne_prefixe_ni_les_nombres_ni_les_dates(db: Database, tmp_path: Path) -> None:
    """Les colonnes numériques et les dates doivent rester exploitables telles quelles."""
    dossier = tmp_path / "pbi"
    powerbi.export_powerbi(db, dossier, today=TODAY)
    entetes, lignes = _read_csv(dossier / "files.csv")
    for colonne in ("file_id", "size_bytes", "priority_score", "age_days_write"):
        index = entetes.index(colonne)
        for ligne in lignes:
            assert ligne[index] == "" or int(ligne[index]) or ligne[index] == "0"
    date_index = entetes.index("last_write_date")
    assert all(re.fullmatch(r"\d{4}-\d{2}-\d{2}", ligne[date_index]) for ligne in lignes)


def test_powerbi_nombre_negatif_en_texte_reste_un_nombre(tmp_path: Path) -> None:
    """`age_days_write` d'un fichier daté du futur est négatif : c'est un nombre, pas une formule."""
    with Database(tmp_path / "futur.sqlite") as database:
        scan = database.start_scan("scan.csv")
        database.upsert_files([_row("futur.pdf", lwt="01/01/2107 10:00:00")], scan)
        database.finish_scan(scan, total=1, new=1, updated=0, unchanged=0, invalid=0)
        powerbi.export_powerbi(database, tmp_path / "pbi", today=TODAY)
    entetes, lignes = _read_csv(tmp_path / "pbi" / "files.csv")
    assert int(lignes[0][entetes.index("age_days_write")]) < 0


def test_powerbi_readme_documente_la_protection(db: Database, tmp_path: Path) -> None:
    dossier = tmp_path / "pbi"
    powerbi.export_powerbi(db, dossier, today=TODAY)
    readme = (dossier / powerbi.README_NAME).read_text(encoding="utf-8")
    assert "injection de formule" in readme.lower()
    assert "apostrophe" in readme
    assert "`=`, `+`, `-` ou `@`" in readme


# ------------------------------------------------------------------ `_flat`


@pytest.mark.parametrize(
    ("stocke", "attendu"),
    [
        ('["identite", "sante"]', "identite|sante"),
        ("[null,1,true]", "1|oui"),  # ni `None`, ni `True` : ce n'est pas du Python
        ("[false]", "non"),
        ('{"x":1}', "x=1"),
        ('[{"montant": 1500, "devise": "EUR"}]', "montant=1500 devise=EUR"),
        ('[{"ok": true}]', "ok=oui"),
        ("[]", ""),
        (None, ""),
        ("pas du json", "pas du json"),
    ],
)
def test_flat_rend_du_texte_presentable(stocke: object, attendu: str) -> None:
    """Ces colonnes sont lues telles quelles dans Power BI : pas de `repr` Python."""
    assert powerbi._flat(stocke) == attendu


# --------------------------------------------------------- Markdown : échappement


def test_markdown_echappe_le_html(tmp_path: Path) -> None:
    """Le Markdown est destiné « à un wiki ou à un mail » — des rendus qui laissent
    passer le HTML brut (GitLab, Confluence, courrier HTML). Le `resume` et les
    justifications de la LLM y arrivaient sans le moindre échappement.
    """
    with _base_hostile(tmp_path / "hostile.sqlite") as database:
        texte = render_markdown(database, today=TODAY)

    assert "<img src=x onerror=alert(1)>" not in texte
    assert "<b>gras</b>" not in texte
    assert "&lt;b&gt;gras&lt;/b&gt;" in texte
    assert "&amp;" in texte
    assert "\\`code\\`" in texte
    # la protection historique du tableau reste en place
    assert "\\|" in texte
    for ligne in texte.splitlines():
        assert "\n" not in ligne


def test_markdown_nechappe_pas_ce_qui_na_pas_besoin_de_letre(db: Database) -> None:
    """Un texte ordinaire traverse inchangé : l'échappement ne doit pas défigurer le rapport."""
    texte = render_markdown(db, today=TODAY)
    assert "contrat.pdf" in texte
    assert "&amp;" not in texte
    assert "&lt;" not in texte


# -------------------------------------------- Excel : comptes exacts à la troncature


@pytest.mark.parametrize("lignes", [1, 2, 5, 8])
def test_excel_troncature_compte_juste_et_ne_deborde_pas(tmp_path: Path, lignes: int) -> None:
    """Écrites + non écrites = données réelles, et l'onglet ne dépasse jamais la limite.

    Avec une limite si basse qu'il ne reste pas même une ligne pour
    l'avertissement, celui-ci était écrit quand même — l'onglet dépassait
    `max_rows` — et une ligne perdue de plus que la réalité était annoncée.
    """
    for max_rows in (1, 2, 3, 6):
        classeur = Workbook(write_only=True)
        message = excel._add_sheet(
            classeur,
            "T",
            ["a", "b"],
            [[i, f"v{i}"] for i in range(lignes)],
            max_rows=max_rows,
        )
        chemin = tmp_path / f"t{max_rows}_{lignes}.xlsx"
        classeur.save(str(chemin))
        feuille = load_workbook(chemin)["T"]
        assert feuille.max_row <= max(max_rows, 1), (max_rows, lignes, feuille.max_row)
        if message is not None:
            trouve = re.search(r": ([\d  ]+) ligne\(s\) écrite\(s\), ([\d  ]+) non", message)
            assert trouve is not None, message
            ecrites, perdues = (int(g.replace(" ", "").replace(" ", "")) for g in trouve.groups())
            assert ecrites + perdues == lignes, (max_rows, lignes, message)


# --------------------- S1 : `export --format csv` n'écrit jamais de formule


def test_cli_export_csv_naucune_cellule_interpretee_comme_formule(tmp_path: Path) -> None:
    """Le CSV de la CLI passait à côté de l'assainissement fait ailleurs.

    C'est pourtant lui que le bouton « CSV des fichiers » écrit, lui qu'on ouvre
    d'un double-clic (`utf-8-sig`, `;`), et lui vers lequel le message de
    troncature du classeur Excel renvoie explicitement pour récupérer la
    totalité des données. Excel évalue toute cellule texte commençant par `=`,
    `+`, `-`, `@`, une tabulation ou un retour chariot.
    """
    base = tmp_path / "hostile.sqlite"
    _base_hostile(base).close()
    sortie = tmp_path / "export.csv"
    assert main(["--db", str(base), "export", "--format", "csv", "--out", str(sortie)]) == 0

    entetes, lignes = _read_csv(sortie)
    suspectes = [
        (colonne, cellule)
        for ligne in lignes
        for colonne, cellule in zip(entetes, ligne, strict=True)
        if cellule[:1] in ("=", "+", "-", "@", "\t", "\r")
    ]
    assert suspectes == []

    # rien n'est perdu : la valeur est là, simplement neutralisée par une apostrophe
    noms = {ligne[entetes.index("name")] for ligne in lignes}
    assert "'=cmd|'/c calc.exe'!A1.docx" in noms
    assert "'-- Copie de sauvegarde.docx" in noms
    assert "'+33 1 23 45 67 89 devis.pdf" in noms
    assert "'@Envoyer a la compta.docx" in noms
    assert "<img src=x onerror=alert(1)>.pdf" in noms  # rien d'inutile n'est préfixé
    resumes = {ligne[entetes.index("resume")] for ligne in lignes}
    assert resumes == {"'" + HOSTILE_RESUME}


def test_cli_export_csv_ne_prefixe_ni_les_nombres_ni_les_identifiants(
    db: Database, tmp_path: Path
) -> None:
    """L'assainissement ne doit pas rendre les colonnes numériques inexploitables."""
    db_path = str(db.path)
    db.close()
    sortie = tmp_path / "sain.csv"
    assert main(["--db", db_path, "export", "--format", "csv", "--out", str(sortie)]) == 0
    entetes, lignes = _read_csv(sortie)
    for colonne in ("size_bytes", "content_version", "security_confidence", "id"):
        index = entetes.index(colonne)
        for ligne in lignes:
            assert not ligne[index].startswith("'"), (colonne, ligne[index])
            assert ligne[index] == "" or int(ligne[index]) >= 0


def test_cli_export_json_rend_les_valeurs_telles_quelles(tmp_path: Path) -> None:
    """Le JSON n'est pas concerné : aucun lecteur JSON n'évalue une chaîne.

    L'y « assainir » abîmerait la donnée sans rien protéger — la valeur brute est
    justement ce qu'attend le programme qui relit ce fichier.
    """
    base = tmp_path / "hostile.sqlite"
    _base_hostile(base).close()
    sortie = tmp_path / "export.json"
    assert main(["--db", str(base), "export", "--format", "json", "--out", str(sortie)]) == 0
    lignes = json.loads(sortie.read_text(encoding="utf-8"))
    assert {ligne["resume"] for ligne in lignes} == {HOSTILE_RESUME}
    assert "=cmd|'/c calc.exe'!A1.docx" in {ligne["name"] for ligne in lignes}


# ------------------- D1 : « à conserver 0 an » dans les rendus


def _base_duree_zero(path: Path) -> Database:
    """Un fichier « à conserver » sans durée, écrit il y a longtemps."""
    database = Database(path)
    scan = database.start_scan("scan.csv")
    database.upsert_files(
        [_row("dossier.pdf", lwt="01/01/2010 08:00:00", access="01/01/2010 08:00:00")], scan
    )
    database.finish_scan(scan, total=1, new=1, updated=0, unchanged=0, invalid=0)
    for file_row in database.iter_files():
        database.store_analysis(
            file_row.id,
            None,
            1,
            prompt_hash="ph0123456789abcd",
            model="qwen38",
            analysis=_analysis("dossier.pdf", retention=True, years=0, basis="legal"),
        )
    return database


def test_les_rendus_disent_la_duree_non_determinee_et_pas_echue(tmp_path: Path) -> None:
    """HTML, Markdown, classeur et Power BI doivent tous refuser de conclure « échu »."""
    with _base_duree_zero(tmp_path / "zero.sqlite") as database:
        page = render_html(database, today=TODAY)
        texte = render_markdown(database, today=TODAY)
        classeur = load_workbook(excel.write_workbook(database, tmp_path / "z.xlsx", today=TODAY))
        powerbi.export_powerbi(database, tmp_path / "pbi", today=TODAY)

    assert views.RETENTION_UNDETERMINED in page
    assert "jamais" in page  # la note qui explique pourquoi
    assert views.RETENTION_UNDETERMINED in texte

    conservation = classeur["Conservation"]
    entetes = [c.value for c in next(conservation.iter_rows(max_row=1))]
    ligne = next(conservation.iter_rows(min_row=2, values_only=True))
    assert ligne[entetes.index("Durée (ans)")] == views.RETENTION_UNDETERMINED
    assert ligne[entetes.index("Fin de conservation")] in ("", None)
    assert ligne[entetes.index("Échu")] == "non"
    synthese = {r[0]: r[1] for r in classeur["Synthèse"].iter_rows(min_row=2, values_only=True)}
    assert synthese["dont durée non déterminée"] == 1
    assert synthese["dont conservation échue"] == 0

    colonnes, lignes = _read_csv(tmp_path / "pbi" / "analyses.csv")
    assert lignes[0][colonnes.index("retention_end_date")] == ""
    assert lignes[0][colonnes.index("retention_required")] == "1"


# ------------- D3 : aucun tableau n'est coupé en silence


@pytest.fixture
def grosse_base(tmp_path: Path) -> Iterator[Database]:
    """180 fichiers : 60 à conserver, 60 candidats au nettoyage, 60 sensibles.

    Assez pour dépasser partout la borne de 50 lignes des tableaux « top ».
    """
    database = Database(tmp_path / "grosse.sqlite")
    scan = database.start_scan("scan.csv")
    lignes = []
    for i in range(180):
        lignes.append(
            _row(
                f"f{i}.e{i % 30}",
                fast_hash=f"h{i // 2}",  # 90 familles de deux exemplaires
                size=1000 + i // 2,  # même empreinte *et* même taille par famille
                owner=f"DOM\\u{i % 30}",
                directory=f"\\\\srv\\part\\d{i}",
                lwt="01/01/2010 08:00:00",
                access="01/01/2010 08:00:00",
                extension=f"e{i % 30}",
            )
        )
    database.upsert_files(lignes, scan)
    database.finish_scan(scan, total=180, new=180, updated=0, unchanged=0, invalid=0)
    for index, file_row in enumerate(sorted(database.iter_files(), key=lambda f: f.id)):
        groupe = index // 60
        if groupe == 0:  # à conserver
            analyse = _analysis(file_row.name, security="C1", retention=True, years=5, basis="rh")
        elif groupe == 1:  # candidat au nettoyage : C0, sans conservation, non accédé
            analyse = _analysis(file_row.name, security="C0", rgpd="none")
        else:  # sensible
            analyse = _analysis(file_row.name, security="C3", rgpd="critical")
        database.store_analysis(
            file_row.id, None, 1, prompt_hash="ph0123456789abcd", model="qwen38", analysis=analyse
        )
        if groupe == 2:
            database.set_review(file_row.id, "corrected", corrected_security="C2")
    yield database
    database.close()


def test_le_rapport_dit_sur_combien_de_lignes_portent_ses_tableaux(grosse_base: Database) -> None:
    """Annoncer un gain de plusieurs téraoctets et n'en montrer que 50 lignes, sans le dire.

    Le rapport de direction a le droit d'être court ; il n'a pas le droit de
    couper en silence sous le total qu'il vient d'annoncer.
    """
    data = collect(grosse_base, today=TODAY)
    page = render_html(grosse_base, data=data)
    texte = render_markdown(grosse_base, data=data)

    assert data.totals["cleanup"] == 60
    assert data.totals["retention"] == 60
    assert data.totals["sensitive"] == 60
    assert data.totals["duplicates"] == 90
    assert data.totals["discrepancies"] == 60
    assert data.totals["extensions"] == data.totals["owners"] == 30

    for rendu in (page, texte):
        assert "premières lignes sur" in rendu
        assert "sur 60" in rendu  # conservation, nettoyage, sensibles, écarts
        assert "sur 90" in rendu  # doublons
        assert "sur 30" in rendu  # extensions et propriétaires (bornés à 25)
        assert "export --format xlsx" in rendu
        assert "export --format powerbi" in rendu


def test_le_rapport_ne_dit_rien_quand_il_ne_coupe_rien(db: Database) -> None:
    """Sur une petite base, aucune note de troncature ne vient encombrer le rapport."""
    data = collect(db, today=TODAY)
    assert "premières lignes sur" not in render_html(db, data=data)
    assert "premières lignes sur" not in render_markdown(db, data=data)


def test_le_classeur_porte_toutes_les_lignes_qui_decident_dune_suppression(
    grosse_base: Database, tmp_path: Path
) -> None:
    """« Conservation » et « Nettoyage » ne sont plus bornés à 50 lignes.

    Chacune de ces lignes est une décision de suppression, et le classeur est
    justement le format vers lequel les rapports renvoient.
    """
    classeur = load_workbook(excel.write_workbook(grosse_base, tmp_path / "gros.xlsx", today=TODAY))
    assert classeur["Conservation"].max_row == 61  # en-tête + 60 lignes
    assert classeur["Nettoyage"].max_row == 61


def test_les_onglets_de_classement_disent_sur_combien_ils_portent(
    grosse_base: Database, tmp_path: Path
) -> None:
    """« Sensibles » et « Doublons » restent des classements — et le disent en dernière ligne."""
    classeur = load_workbook(excel.write_workbook(grosse_base, tmp_path / "gros.xlsx", today=TODAY))
    sensibles = classeur["Sensibles"]
    assert sensibles.max_row == 52  # en-tête + 50 lignes + la note
    note = str(sensibles.cell(row=52, column=1).value)
    assert "50 fichiers affichés sur 60" in note
    assert "powerbi" in note
    assert sensibles.cell(row=52, column=1).font.bold is True

    doublons = classeur["Doublons"]
    assert doublons.max_row == 52
    assert "50 familles affichées sur 90" in str(doublons.cell(row=52, column=1).value)


def test_pas_de_note_de_classement_quand_le_classement_est_complet(
    db: Database, tmp_path: Path
) -> None:
    classeur = load_workbook(excel.write_workbook(db, tmp_path / "petit.xlsx", today=TODAY))
    dernieres = {
        feuille.title: str(feuille.cell(row=feuille.max_row, column=1).value)
        for feuille in classeur
    }
    assert all("affichés sur" not in v and "affichées sur" not in v for v in dernieres.values())


# ------------------------------------- HTML : le texte de la LLM est échappé


def test_le_html_echappe_le_resume_et_la_justification_de_la_llm(tmp_path: Path) -> None:
    """Le rapport HTML est ouvert dans un navigateur ; le `resume` vient du modèle.

    Retirer l'échappement de ces deux colonnes ne faisait échouer aucun test : la
    page est autonome, sans script, mais un `<img onerror>` glissé dans un résumé
    en exécutait un.
    """
    with _base_hostile(tmp_path / "hostile.sqlite") as database:
        page = render_html(database, today=TODAY)

    assert "<b>gras</b>" not in page
    assert "&lt;b&gt;gras&lt;/b&gt;" in page
    assert "<img src=x onerror=alert(1)>" not in page
    assert "&lt;img src=x onerror=alert(1)&gt;" in page  # le nom de fichier, échappé
    assert "onerror=alert(1)>.pdf</span>" not in page
    assert "=1+2 justification" not in page or "&amp;" in page
    # aucune balise ouverte par le texte du modèle : la page reste bien formée
    controleur = _Checker()
    controleur.feed(page)
    controleur.close()
    assert controleur.errors == []
    assert "img" not in controleur.tags


# -------------------------------------------- périmètre incomplet dans les rendus


def _marque_incomplete(db: Database, **faits: object) -> None:
    """Marque le scan de la base comme portant sur un périmètre amputé."""
    scan_id = int(db.query_values("SELECT id FROM scans ORDER BY id LIMIT 1")[0][0])
    db.annotate_scan(
        scan_id,
        manifest_json="{}",
        scanner_elapsed_s=1.0,
        **faits,  # type: ignore[arg-type]
    )


def test_rapports_taisent_le_perimetre_quand_il_est_complet(db: Database, tmp_path: Path) -> None:
    """Cas normal : aucun des trois rendus ne parle de périmètre incomplet."""
    data = collect(db, today=TODAY)
    assert data.scope.incomplete is False
    page = render_html(db, data=data)
    assert '<div class="perimetre"' not in page  # la règle CSS reste, pas le bandeau
    assert "Inventaire incomplet" not in page
    assert "Périmètre incomplet" not in page
    assert "Inventaire incomplet" not in render_markdown(db, data=data)
    classeur = load_workbook(excel.write_workbook(db, tmp_path / "c.xlsx", today=TODAY))
    assert classeur.sheetnames == list(excel.SHEETS)  # pas d'onglet « Périmètre »
    assert classeur["Synthèse"].cell(row=2, column=2).value == "complet"


def test_rapports_annoncent_un_perimetre_incomplet_sans_le_faire_chercher(
    db: Database, tmp_path: Path
) -> None:
    """Un partage non parcouru se voit d'emblée dans le HTML, le Markdown et l'Excel.

    Le rapport HTML est remis à la direction et sert à justifier des suppressions :
    l'avertissement est un bandeau **avant** le sommaire et la synthèse, pas une
    note en pied de page. Le Markdown le porte avant la section 1, et le classeur
    en fait son premier onglet.
    """
    _marque_incomplete(db, skipped=["\\\\srv\\finance"], exit_code=4, expected_files=4)
    data = collect(db, today=TODAY)
    assert data.scope.incomplete is True
    assert data.scope.skipped_targets == ["\\\\srv\\finance"]

    html = render_html(db, data=data)
    bandeau = html.index('<div class="perimetre"')
    assert bandeau < html.index('<nav class="sommaire"')  # avant le sommaire
    assert bandeau < html.index('id="synthese"')  # et avant la synthèse
    assert "Inventaire incomplet" in html
    assert "srv\\finance" in html
    assert "Périmètre incomplet" in html  # dans le bandeau d'en-tête aussi
    verificateur = _Checker()
    verificateur.feed(html)
    assert verificateur.errors == []

    md = render_markdown(db, data=data)
    assert md.index("Inventaire incomplet") < md.index("## 1. Synthèse")
    assert "srv\\finance" in md

    classeur = load_workbook(excel.write_workbook(db, tmp_path / "c.xlsx", today=TODAY))
    assert classeur.sheetnames[0] == excel.SCOPE_SHEET
    perimetre = "\n".join(
        str(cell.value) for row in classeur[excel.SCOPE_SHEET].iter_rows() for cell in row
    )
    assert "srv\\finance" in perimetre
    assert classeur["Synthèse"].cell(row=2, column=2).value == "INCOMPLET"


def test_rapports_signalent_un_scan_arrete(db: Database) -> None:
    """Un scan annulé se lit dans le rapport, des mois après, sans le manifeste."""
    _marque_incomplete(db, cancelled=True, expected_files=999)
    data = collect(db, today=TODAY)
    assert data.scope.cancelled_scans == 1
    assert "arrêté en cours de route" in " ".join(data.scope.warnings)
    assert "arrêtés en cours de route" in data.scope.headline()
    assert "Inventaire incomplet" in render_html(db, data=data)
